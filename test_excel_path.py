import os
from openpyxl import load_workbook

# Test Excel file path
EXCEL_FILE_PATH = r"D:\Anant\Youtube\ValueProITGyan\YouTubeVideosList.xlsx"
SHEET_NAME = "Shorts_Automation"

print(f"🔍 Testing Excel file access...")
print(f"📁 File path: {EXCEL_FILE_PATH}")
print(f"📄 File exists: {os.path.exists(EXCEL_FILE_PATH)}")

if os.path.exists(EXCEL_FILE_PATH):
    try:
        workbook = load_workbook(EXCEL_FILE_PATH, data_only=True)
        print(f"✅ Excel file loaded successfully")
        print(f"📋 Available sheets: {workbook.sheetnames}")
        
        if SHEET_NAME in workbook.sheetnames:
            sheet = workbook[SHEET_NAME]
            print(f"✅ Sheet '{SHEET_NAME}' found")
            
            # Test reading specific cells
            cells_to_read = ['C2', 'C3', 'C9', 'C10', 'C11', 'C12']
            for cell in cells_to_read:
                value = sheet[cell].value
                print(f"   📝 {cell}: {value}")
                
        else:
            print(f"❌ Sheet '{SHEET_NAME}' not found")
            
        workbook.close()
        
    except Exception as e:
        print(f"❌ Error accessing Excel file: {e}")
else:
    print(f"❌ Excel file not found at specified path")
    print(f"🔍 Please verify the file path is correct")
