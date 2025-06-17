# Test script to verify Excel reading functionality
import os
from openpyxl import load_workbook

EXCEL_FILE_PATH = r"D:\Anant\Youtube\ValueProITGyan\YouTubeVideosList.xlsx"
SHEET_NAME = "Shorts_Automation"

def test_excel_read():
    """Test reading Excel file"""
    try:
        print(f"📖 Testing Excel file: {EXCEL_FILE_PATH}")
        print(f"📁 File exists: {os.path.exists(EXCEL_FILE_PATH)}")
        
        if not os.path.exists(EXCEL_FILE_PATH):
            print("❌ Excel file not found!")
            return False
            
        workbook = load_workbook(EXCEL_FILE_PATH)
        print(f"📊 Available sheets: {workbook.sheetnames}")
        
        if SHEET_NAME not in workbook.sheetnames:
            print(f"❌ Sheet '{SHEET_NAME}' not found!")
            workbook.close()
            return False
            
        sheet = workbook[SHEET_NAME]
        
        # Read data from specified cells
        cells_to_read = ['C2', 'C3', 'C9', 'C10', 'C11', 'C12']
        data = {}
        
        print("\n📋 Reading cell data:")
        for cell in cells_to_read:
            value = sheet[cell].value
            data[cell] = value
            print(f"   {cell}: {repr(value)}")
        
        workbook.close()
        print("\n✅ Excel reading test completed successfully!")
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

if __name__ == "__main__":
    test_excel_read()
