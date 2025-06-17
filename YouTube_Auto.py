import requests
import json
import os
from dotenv import load_dotenv
from openpyxl import load_workbook

# Load environment variables from .env file
load_dotenv()

# Configuration
EXCEL_FILE_PATH = r"D:\Anant\Youtube\ValueProITGyan\YouTubeVideosList.xlsx"
SHEET_NAME = "Shorts_Automation"

def get_ai_response(prompt, model="deepseek/deepseek-chat:free"):
    """Get AI response from OpenRouter API"""
    api_key = os.getenv('OPENROUTER_API_KEY', 'your-api-key-here')
    
    if api_key == 'your-api-key-here':
        print("❌ Error: API key not found. Check your .env file.")
        return None
    
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "HTTP-Referer": "https://localhost",
        "X-Title": "YouTube AI Chat",
    }
    
    data = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}]
    }
    
    try:
        print(f"🤖 Making API request for prompt: {prompt[:50]}...")
        response = requests.post(
            url="https://openrouter.ai/api/v1/chat/completions",
            headers=headers,
            data=json.dumps(data)
        )
        
        if response.status_code == 200:
            result = response.json()
            if 'choices' in result and len(result['choices']) > 0:
                return result['choices'][0]['message']['content']
            else:
                print("❌ No response content found in API response")
                return None
        else:
            print(f"❌ API Error: HTTP {response.status_code}")
            if response.status_code == 401:
                print("🔑 Authentication Error: Invalid API key")
            print("Response:", response.text)
            return None
            
    except Exception as e:
        print(f"❌ Request failed: {e}")
        return None

def read_excel_data():
    """Read data from Excel file"""
    try:
        print(f"📖 Reading Excel file: {EXCEL_FILE_PATH}")
        print(f"📁 File exists: {os.path.exists(EXCEL_FILE_PATH)}")
        
        if not os.path.exists(EXCEL_FILE_PATH):
            print("❌ Excel file not found!")
            return None
            
        workbook = load_workbook(EXCEL_FILE_PATH)
        print(f"📊 Available sheets: {workbook.sheetnames}")
        
        if SHEET_NAME not in workbook.sheetnames:
            print(f"❌ Sheet '{SHEET_NAME}' not found in workbook")
            print(f"Available sheets: {workbook.sheetnames}")
            workbook.close()
            return None
            
        sheet = workbook[SHEET_NAME]
        
        # Read data from specified cells
        data = {
            'C2': sheet['C2'].value,
            'C3': sheet['C3'].value,
            'C9': sheet['C9'].value,
            'C10': sheet['C10'].value,
            'C11': sheet['C11'].value,
            'C12': sheet['C12'].value
        }
        
        print("✅ Excel data read successfully:")
        for cell, value in data.items():
            print(f"   {cell}: {repr(value)}")
        
        workbook.close()
        return data
        
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return None

def write_to_excel(cell, value):
    """Write data to Excel cell"""
    try:
        workbook = load_workbook(EXCEL_FILE_PATH)
        sheet = workbook[SHEET_NAME]
        sheet[cell] = value
        workbook.save(EXCEL_FILE_PATH)
        workbook.close()
        print(f"✅ Written to Excel cell {cell}")
        return True
    except Exception as e:
        print(f"❌ Error writing to Excel: {e}")
        return False

def write_to_text_file(filename, content):
    """Write content to text file"""
    try:
        filepath = os.path.join(os.path.dirname(EXCEL_FILE_PATH), f"{filename}.txt")
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(content)
        print(f"✅ Written to text file: {filename}.txt")
        return True
    except Exception as e:
        print(f"❌ Error writing to text file: {e}")
        return False

def main():
    """Main automation workflow - AUTO MODE (no user prompts)"""
    print("🚀 Starting YouTube AI Chat Excel Automation - AUTO MODE")
    print("=" * 60)
    
    # Step 1: Read Excel data
    excel_data = read_excel_data()
    if not excel_data:
        print("❌ Failed to read Excel data. Exiting.")
        return
    
    print("\n📋 Excel data loaded successfully. Processing with AI...")
    
    # Step 2: Process C2 content for B4
    if excel_data['C2']:
        print(f"\n🔄 Step 1: Processing C2 content for B4...")
        prompt_c2 = f"Generate a YouTube video title based on this content: '{excel_data['C2']}'. Make it engaging and clickable for B4 cell."
        
        response_b4 = get_ai_response(prompt_c2)
        if response_b4:
            print(f"📝 AI Response for B4:\n{response_b4[:200]}...")
            write_to_excel('B4', response_b4)
    
    # Step 3: Process C3 content for B6
    if excel_data['C3']:
        print(f"\n🔄 Step 2: Processing C3 content for B6...")
        prompt_c3 = f"Generate a YouTube video description based on this content: '{excel_data['C3']}'. Make it SEO optimized for B6 cell."
        
        response_b6 = get_ai_response(prompt_c3)
        if response_b6:
            print(f"📝 AI Response for B6:\n{response_b6[:200]}...")
            write_to_excel('B6', response_b6)
    
    # Step 4: Process C9 content for B9 and ShortEng_AT file
    if excel_data['C9']:
        print(f"\n🔄 Step 3: Processing C9 content for B9 and ShortEng_AT...")
        prompt_c9 = f"Create an English YouTube Short script based on this content: '{excel_data['C9']}'. Keep it under 60 seconds and engaging."
        
        response_b9 = get_ai_response(prompt_c9)
        if response_b9:
            print(f"📝 AI Response for B9/ShortEng_AT:\n{response_b9[:200]}...")
            write_to_excel('B9', response_b9)
            write_to_text_file('ShortEng_AT', response_b9)
    
    # Step 5: Process C10 content for B10 and ShortHindi_AT file
    if excel_data['C10']:
        print(f"\n🔄 Step 4: Processing C10 content for B10 and ShortHindi_AT...")
        prompt_c10 = f"Create a Hindi YouTube Short script based on this content: '{excel_data['C10']}'. Keep it under 60 seconds and engaging. Respond in Hindi."
        
        response_b10 = get_ai_response(prompt_c10)
        if response_b10:
            print(f"📝 AI Response for B10/ShortHindi_AT:\n{response_b10[:200]}...")
            write_to_excel('B10', response_b10)
            write_to_text_file('ShortHindi_AT', response_b10)
    
    print("\n🎉 Excel automation workflow completed!")
    print("=" * 60)
    print("📋 Summary:")
    print("   ✅ Read data from Excel cells C2, C3, C9, C10, C11, C12")
    print("   ✅ Generated AI responses and saved to:")
    print("      - B4 (from C2 content)")
    print("      - B6 (from C3 content)")
    print("      - B9 (from C9 content)")
    print("      - B10 (from C10 content)")
    print("   ✅ Created text files:")
    print("      - ShortEng_AT.txt")
    print("      - ShortHindi_AT.txt")

if __name__ == "__main__":
    main()
